from django.http import HttpResponse
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import JsonResponse
import json
from user.models import Student
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Custom 404 error view
def custom_404_view(request, exception):
    return render(request, '404.html', status=404)

# Custom 500 error view
def custom_500_view(request):
    return render(request, '500.html', status=500)

# College portal view
def college_portal(request):
    # Example role checking logic
    if request.user.is_authenticated:
        # Check if the user belongs to the 'Student' group
        if request.user.groups.filter(name='Student').exists():
            return redirect('student_dashboard') 
        # Check if the user belongs to the 'Teacher' group
        elif request.user.groups.filter(name='Teacher').exists():
            return redirect('teacher_dashboard') 
        # Check if the user is a superuser (admin)
        elif request.user.is_superuser:
            return redirect('admin_dashboard') 
    # If the user is not authenticated or doesn't belong to any group, show a landing page or generic content
    context = {
        # SEO-related content
        'description': '🎓 Become an Academic Achiever at IIH College! 🎓\n\n'
                       'Are you a Grade 10 student from a public school graduating in 2024? IIH College offers an incredible opportunity for you to pursue your Senior High School education with ZERO PAYMENT through the DepEd SHS Voucher Program!\n\n'
                       'Here’s what you can expect when you enroll with us:\n\n'
                       '✅ No Entrance Exam\n'
                       '✅ No Maintaining Grade Requirement\n'
                       '✅ No Registration Fees\n'
                       '✅ No Examination Fees\n'
                       '✅ No Monthly Fees\n'
                       '✅ No Hidden Charges\n\n'
                       'PLUS, enjoy these amazing FREE items:\n'
                       '🎽 FREE Daily Uniform (1 set)\n'
                       '🎽 FREE PE Uniform (1 set)\n'
                       '📛 FREE ID & Lace\n\n'
                       'Whether you choose Face-to-Face or Full Modular Learning Modality, your education is completely covered!\n\n'
                       "Don't miss this chance to start your journey towards academic excellence. ENROLL NOW!",
        
        # Keywords for SEO (separated by commas)
        'keywords': 'Academic Achiever, Senior High School Enrollment, Zero Payment Education, Public School Graduates, DepEd SHS Voucher Program, Free School Uniforms, Enroll Now, iihcollege',
        
        # Open Graph meta description
        'og_description': 'Enroll at IIH College for Senior High School with ZERO PAYMENT for Grade 10 public school graduates in 2024! Enjoy a hassle-free experience with no entrance exams, no fees, and no hidden charges.',
        
        # Canonical URL for the page (to avoid duplicate content)
        'canonical_url': 'https://www.yourwebsite.com/college-portal',
   
        # Page title and description for SEO
        'page_title': 'IIH College Portal - Zero Payment Enrollment for Senior High',
        'page_description': 'IIH College offers senior high school enrollment with zero payment for grade 10 public school graduates in 2024.',
        'page_url': 'https://www.yourwebsite.com/college-port2l',
        'year': '2024',
    }
    return render(request, 'home.html', context)

def denied(request):
        context = {
        'header_title': 'Erollment Form',
        'page_title': 'Denied Access',
    }
        return render(request, 'denied.html', context)

def about(request):
        context = {
        'header_title': 'About',
        'page_title': 'iiH College Portal - About',
    }
        return render(request, 'about-us.html', context)

def soon(request):
    context = {
        'header_title': 'Soon',
        'page_title': 'IIH College Portal - Soon',
    }
    return render(request, 'soon.html', context)

def privacy(request):
    context = {
        'header_title': 'Privacy Policy',
        'page_title': 'IIH College Portal - Privacy Policy',
    }
    return render(request, 'privacy_policy.html', context)

def terms(request):
    context = {
        'header_title': 'Terms and Conditions',
        'page_title': 'IIH College Portal - Terms',

    }
    return render(request, 'terms_of_service.html', context)

def success(request):
    context = {
        'header_title': 'Terms and Conditions',
        'page_title': 'IIH College Portal - Success',

    }
    return render(request, 'success.html', context)

def student_list(request):
    # Fetching all approved students
    students = Student.objects.filter(status=Student.APPROVED)

    # Initialize the live_excel_url variable
    live_excel_url = ""

    if request.method == 'POST':
        live_excel_url = request.POST.get('live_excel_url', "")

    # URL for the Excel file (you can use this as a fallback or pre-configure a default one)
    excel_file_url = '/media/uploads/Book.xlsx'
    
    context = {
        'page_title': 'IIH College Portal - Student List',
        'header_title': 'Student List',
        'students': students,
        'excel_file_url': excel_file_url,
        'live_excel_url': live_excel_url,  # The dynamic live Excel URL entered by the user
    }
    
    return render(request, 'student_list.html', context)

def export_students_to_excel(request):
    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student List"
    
    # Add the header row
    headers = ["Student ID", "Full Name", "Email", "Phone Number", "Strand", "Status"]
    ws.append(headers)
    
    # Apply style to the header row
    header_font = Font(bold=True, color="FFFFFF")  # White font, bold
    header_fill = openpyxl.styles.PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green background
    header_alignment = Alignment(horizontal="center", vertical="center")  # Center-align text
    
    # Apply header style and adjust column width
    for cell in ws[1]:  # Apply styles to the first row (header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(bottom=Side(style="thin"))  # Border under the header
    
    # Adjust column widths for better visibility
    column_widths = {
        'A': 15,  # Student ID
        'B': 25,  # Full Name
        'C': 30,  # Email
        'D': 20,  # Phone Number
        'E': 60,  # Strand
        'F': 15   # Status
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # Get all approved students
    students = Student.objects.filter(status=Student.APPROVED)
    
    # Add student data to the sheet with styling
    row_num = 2  # Start from row 2 (after the header)
    for student in students:
        row = [
            student.student_id,
            f"{student.first_name} {student.last_name}",
            student.email,
            student.phone,  # Assuming phone is stored as a string
            student.get_strand_display(),  # Display the choice value of strand
            student.get_status_display(),  # Display the choice value of status
        ]
        ws.append(row)

        # Apply styles to the data rows
        for col_num, cell in enumerate(ws[row_num], start=1):
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center-align text
            cell.border = Border(bottom=Side(style="thin"))  # Thin bottom border for each cell
        row_num += 1

    # Set up the HTTP response to return the Excel file as an attachment
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=student_list.xlsx"
    
    # Save the workbook to the response
    wb.save(response)
    
    return response

